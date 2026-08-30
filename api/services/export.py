"""Scope-aware exports — every export respects whatever filters the caller is
currently viewing (platform / module / sub_module / severity / phase / owner)."""

import csv
import io
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from models.defect import Defect
from models.test_case import TestCase

BRAND_INDIGO = "5B5BF6"
BRAND_TEAL = "00C9A7"
SEVERITY_FILL = {
    "Critical": "F8D7DA",
    "High": "FDEBD0",
    "Medium": "FFF3CD",
    "Low": "D6EAF8",
    "TBC": "E5E7EB",
}


async def get_filtered_defects(
    session: AsyncSession, platform=None, module=None, sub_module=None,
    severity=None, owner=None, state=None,
) -> list[Defect]:
    q = select(Defect)
    if platform:
        q = q.where(Defect.platform == platform)
    if module:
        q = q.where(Defect.module == module)
    if sub_module:
        q = q.where(Defect.sub_module == sub_module)
    if severity:
        q = q.where(Defect.severity == severity)
    if owner:
        q = q.where(Defect.owner_vendor == owner)
    if state:
        q = q.where(Defect.state == state)
    q = q.order_by(Defect.raised_date.desc().nulls_last())
    return list((await session.execute(q)).scalars().all())


async def get_filtered_tests(
    session: AsyncSession, platform=None, module=None, sub_module=None, phase=None,
) -> list[TestCase]:
    q = select(TestCase)
    if platform:
        q = q.where(TestCase.platform == platform)
    if module:
        q = q.where(TestCase.module == module)
    if sub_module:
        q = q.where(TestCase.sub_module == sub_module)
    if phase:
        q = q.where(TestCase.phase == phase)
    q = q.order_by(TestCase.executed_date.desc().nulls_last())
    return list((await session.execute(q)).scalars().all())


def _aging(d: Defect) -> int | None:
    if not d.raised_date:
        return None
    return (date.today() - d.raised_date).days


DEFECT_HEADERS = [
    "External ID", "Source", "Platform", "Module", "Sub-module", "Title", "Severity",
    "State", "Owner Vendor", "Assignee Email", "Raised Date", "ETA", "Aging (days)",
    "Resolved Date", "Reopened", "Remarks",
]

TEST_HEADERS = [
    "External ID", "Source", "Platform", "Module", "Sub-module", "Title", "Status",
    "Phase", "Executed Date",
]


def _defect_row(d: Defect) -> list:
    return [
        d.external_id, d.source.value, d.platform, d.module, d.sub_module, d.title,
        d.severity.value, d.state, d.owner_vendor or "", d.assignee_email or "",
        d.raised_date.isoformat() if d.raised_date else "",
        d.eta.isoformat() if d.eta else "",
        _aging(d) if _aging(d) is not None else "",
        d.resolved_date.isoformat() if d.resolved_date else "",
        "Yes" if d.is_reopen else "No",
        d.remarks or "",
    ]


def _test_row(t: TestCase) -> list:
    return [
        t.external_id, t.source.value, t.platform, t.module, t.sub_module, t.title,
        t.status.value, t.phase.value if t.phase else "",
        t.executed_date.isoformat() if t.executed_date else "",
    ]


# ── CSV ──────────────────────────────────────────────────────────────────────

def defects_to_csv(defects: list[Defect]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(DEFECT_HEADERS)
    for d in defects:
        writer.writerow(_defect_row(d))
    return buf.getvalue().encode("utf-8")


def tests_to_csv(tests: list[TestCase]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(TEST_HEADERS)
    for t in tests:
        writer.writerow(_test_row(t))
    return buf.getvalue().encode("utf-8")


# ── Excel (openpyxl) ─────────────────────────────────────────────────────────

def _style_header(ws, headers: list[str]):
    fill = PatternFill(start_color=BRAND_INDIGO, end_color=BRAND_INDIGO, fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def defects_to_excel(defects: list[Defect]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Defects"
    _style_header(ws, DEFECT_HEADERS)

    severity_col = DEFECT_HEADERS.index("Severity") + 1
    remarks_col = DEFECT_HEADERS.index("Remarks") + 1

    for r, d in enumerate(defects, start=2):
        for c, value in enumerate(_defect_row(d), start=1):
            cell = ws.cell(row=r, column=c, value=value)
            if c == remarks_col:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        sev_cell = ws.cell(row=r, column=severity_col)
        fill_color = SEVERITY_FILL.get(d.severity.value)
        if fill_color:
            sev_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    widths = [14, 8, 16, 16, 16, 40, 10, 14, 16, 26, 12, 12, 12, 14, 10, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def tests_to_excel(tests: list[TestCase]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tests"
    _style_header(ws, TEST_HEADERS)

    status_col = TEST_HEADERS.index("Status") + 1
    status_fill = {"Passed": "D4EDDA", "Failed": "F8D7DA", "Blocked": "FFF3CD",
                   "Not Run": "E5E7EB", "In Progress": "D6EAF8"}

    for r, t in enumerate(tests, start=2):
        for c, value in enumerate(_test_row(t), start=1):
            ws.cell(row=r, column=c, value=value)
        status_cell = ws.cell(row=r, column=status_col)
        fill_color = status_fill.get(t.status.value)
        if fill_color:
            status_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    widths = [14, 8, 16, 16, 16, 40, 14, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF (reportlab) ──────────────────────────────────────────────────────────

def _branded_doc(buf, title: str):
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=14 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Heading1"], textColor=colors.HexColor(f"#{BRAND_INDIGO}"))
    subtitle_style = ParagraphStyle("Subtitle", parent=styles["Normal"], textColor=colors.grey)
    story = [
        Paragraph("QA Command", title_style),
        Paragraph(title, subtitle_style),
        Paragraph(f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", subtitle_style),
        Spacer(1, 8),
    ]
    return doc, story


def _table_style(header_row=0):
    return TableStyle([
        ("BACKGROUND", (0, header_row), (-1, header_row), colors.HexColor(f"#{BRAND_INDIGO}")),
        ("TEXTCOLOR", (0, header_row), (-1, header_row), colors.white),
        ("FONTNAME", (0, header_row), (-1, header_row), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("ROWBACKGROUNDS", (0, header_row + 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FC")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ])


def defects_to_pdf(defects: list[Defect], scope_label: str = "All defects") -> bytes:
    buf = io.BytesIO()
    doc, story = _branded_doc(buf, f"Defects export — {scope_label} ({len(defects)} rows)")
    headers = [h for h in DEFECT_HEADERS if h != "Remarks"]
    data = [headers] + [[str(v) for v in _defect_row(d)[:-1]] for d in defects]
    table = Table(data, repeatRows=1)
    table.setStyle(_table_style())
    story.append(table)
    doc.build(story)
    return buf.getvalue()


def tests_to_pdf(tests: list[TestCase], scope_label: str = "All tests") -> bytes:
    buf = io.BytesIO()
    doc, story = _branded_doc(buf, f"Tests export — {scope_label} ({len(tests)} rows)")
    data = [TEST_HEADERS] + [[str(v) for v in _test_row(t)] for t in tests]
    table = Table(data, repeatRows=1)
    table.setStyle(_table_style())
    doc.build(story + [table])
    return buf.getvalue()


def snapshot_pdf(tree: dict) -> bytes:
    """Branded one-pager of the overall + per-platform health — used as the daily
    email's PDF attachment."""
    buf = io.BytesIO()
    doc, story = _branded_doc(buf, "Daily snapshot")
    overall = tree["metrics"]
    styles = getSampleStyleSheet()
    story.append(Paragraph(
        f"Execution {overall['execution_pct']}% · Pass rate {overall['pass_rate']}% · "
        f"Open defects {overall['open_defects']} (Critical {overall['open_by_severity'].get('Critical', 0)}) · "
        f"Aging&gt;21d {overall['aging_gt21']} · Avg resolution {overall['avg_resolution_days'] or '—'}d",
        styles["Normal"],
    ))
    story.append(Spacer(1, 10))

    rows = [["Platform", "Health", "Execution %", "Open Defects", "Critical", "Aging>21"]]
    for name, p in tree.get("platforms", {}).items():
        m = p["metrics"]
        rows.append([name, m["health"].upper(), m["execution_pct"], m["open_defects"],
                     m["open_by_severity"].get("Critical", 0), m["aging_gt21"]])
    table = Table(rows, repeatRows=1)
    table.setStyle(_table_style())
    story.append(table)
    doc.build(story)
    return buf.getvalue()
